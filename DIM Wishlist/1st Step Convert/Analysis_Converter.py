import json
import os
from openpyxl import load_workbook

def excel_to_json():
    # Define relative paths
    script_dir = os.path.dirname(__file__)
    excel_file = os.path.join(script_dir, 'Destiny 2_ Endgame Analysis.xlsx')
    output_dir = os.path.join(script_dir, 'Json')

    # Check for Excel file
    if not os.path.exists(excel_file):
        return

    # Load Excel file
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    weapon_sheets = ['Autos', 'Bows', 'HCs', 'Pulses', 'Scouts', 'Sidearms', 'SMGs', 'BGLs', 'Fusions', 'Glaives', 'Shotguns', 'Snipers', 'HGLs', 'LFRs', 'LMGs', 'Rockets', 'Swords', 'Other', 'Rocket Sidearms', 'Traces']

    for sheet_name in weapon_sheets:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        weapons = []

        # Get headers from row 2
        headers = [cell.value.lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=2, max_row=2))]
        name_idx = headers.index('name') if 'name' in headers else None
        col1_idx = headers.index('column 1') if 'column 1' in headers else None
        col2_idx = headers.index('column 2') if 'column 2' in headers else None
        origin_idx = headers.index('origin trait') if 'origin trait' in headers else None

        # Find the last column with data in row 2
        last_col_idx = None
        for idx, cell in enumerate(headers):
            if cell.strip():
                last_col_idx = idx

        # Process data rows (starting from row 3)
        for row in ws.iter_rows(min_row=3):
            if name_idx is None or not row[name_idx].value:
                continue

            # Get tier rating from last column
            tier = str(row[last_col_idx].value).strip().upper() if last_col_idx is not None and row[last_col_idx].value else ''
            if sheet_name != 'Other' and tier not in ['A', 'S']:
                continue  # Skip non-A/S tier weapons, except for 'Other'

            weapon = {}
            # Handle multi-line names
            name = str(row[name_idx].value).strip()
            if '\n' in name:
                parts = name.split('\n')
                name = parts[0].strip()
                for part in parts[1:]:
                    if 'version' in part.lower():
                        name += ' ' + part.strip()
            weapon['name'] = name

            weapon['column1'] = str(row[col1_idx].value).replace('\n', ', ').strip() if col1_idx is not None and row[col1_idx].value else 'None'
            weapon['column2'] = str(row[col2_idx].value).replace('\n', ', ').strip() if col2_idx is not None and row[col2_idx].value else 'None'
            weapon['originTrait'] = str(row[origin_idx].value).strip() if origin_idx is not None and row[origin_idx].value else 'None'

            weapons.append(weapon)

        # Save to category-specific JSON file
        if weapons:
            output_filename = f"{sheet_name}.json"
            output_path = os.path.join(output_dir, output_filename)
            os.makedirs(output_dir, exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump({"Weapons": weapons}, f, indent=4)

if __name__ == "__main__":
    excel_to_json()